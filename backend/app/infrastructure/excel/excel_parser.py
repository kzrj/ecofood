"""
Инфраструктурный слой: знает про openpyxl, не знает про домен.
Принимает сырые байты → парсит, группирует по типам товаров.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Any

import openpyxl

log = logging.getLogger(__name__)

_TARGET_SHEET_PREFIX = "завод план"
_MAX_COLUMNS = 6
_MAX_PARSE_COLUMN_1BASED = 24  # X

# Полный список колонок (нужен для детектирования типов и товаров по индексам).
# Колонки 0 (замес) и 3 (артикул) используются только внутри парсера и в выходной dict не попадают.
_COLUMN_NAMES = ["замес", "замес на 100", "замес на 150", "артикул", "потребность", "наименование"]

# Индексы и имена колонок, которые попадают в итоговый dict.
_OUTPUT_COLUMNS: dict[int, str] = {1: "замес на 100", 2: "замес на 150", 4: "потребность", 5: "наименование"}
_WEEKDAYS = {"пн", "вт", "ср", "чт", "пт", "сб", "вс"}

# Известные типы продуктов (нижний регистр, более длинные — раньше).
_KNOWN_TYPES: list[str] = [
    "варено-копченные",
    "варено-копченые",
    "вареные",
    "ветчины",
    "сосиски-сардельки",
    "сосиски",
    "сардельки",
    "холодцы",
    "деликатес",
    "пф для сп",
]


def _select_target_sheet(workbook: openpyxl.Workbook):
    log.info("Листы в книге: %s", [s.title for s in workbook.worksheets])
    for sheet in workbook.worksheets:
        if sheet.title.strip().lower().startswith(_TARGET_SHEET_PREFIX):
            log.info("Выбран лист: '%s'", sheet.title)
            return sheet
    raise ValueError("Лист с названием, начинающимся на 'Завод план', не найден")


def _detect_type(row_slice: tuple) -> str | None:
    """
    Строка является заголовком типа, если:
      - колонка «замес» (index 0) == None  (у товаров там число 1 или 3)
      - колонка «наименование» (index 5) содержит название типа
    Возвращает исходное значение из ячейки (stripped) или None.
    """
    if row_slice[0] is not None:
        return None
    last = row_slice[5] if len(row_slice) > 5 else None
    if last is None:
        return None
    val = str(last).strip().lower()
    for known in _KNOWN_TYPES:
        if val.startswith(known):
            return str(last).strip()
    return None


def _is_product_row(row_slice: tuple) -> bool:
    """Строка товара: колонка «замес» — число (1 или 3)."""
    return isinstance(row_slice[0], (int, float))


def _normalize_weekday(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip().lower().replace(".", "")
    if s in _WEEKDAYS:
        return s.upper()
    return None


def _format_ddmm(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%d.%m")
    if isinstance(value, date):
        return value.strftime("%d.%m")
    if value is None:
        return None
    s = str(value).strip()
    m = re.search(r"(\d{1,2})[./](\d{1,2})", s)
    if not m:
        return None
    d, mth = int(m.group(1)), int(m.group(2))
    if not (1 <= d <= 31 and 1 <= mth <= 12):
        return None
    return f"{d:02d}.{mth:02d}"


def _detect_day_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[int, str]:
    """Возвращает map: индекс_в_row (0-based) -> ключ дня 'ПН 25.03'."""
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        return {}
    row1, row2 = rows[0], rows[1]
    max_len = min(max(len(row1), len(row2)), _MAX_PARSE_COLUMN_1BASED)

    result: dict[int, str] = {}
    for idx in range(_MAX_COLUMNS, max_len):
        wd = _normalize_weekday(row1[idx] if idx < len(row1) else None)
        ddmm = _format_ddmm(row2[idx] if idx < len(row2) else None)
        if wd and ddmm:
            result[idx] = f"{wd} {ddmm}"
    log.info("Дневные колонки: %s", result)
    return result


def parse_excel_bytes(data: bytes) -> list[dict[str, Any]]:
    """
    Читает лист 'Завод план …', первые 6 колонок.
    Возвращает плоский список dict (без группировки).
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = _select_target_sheet(wb)

    result: list[dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        row_slice = row[:_MAX_COLUMNS]
        if all(v is None for v in row_slice):
            continue
        result.append({name: row_slice[i] for i, name in _OUTPUT_COLUMNS.items()})

    log.info("Итого строк: %d", len(result))
    wb.close()
    return result


def group_rows_by_type(data: bytes) -> dict[str, list[dict[str, Any]]]:
    """
    Читает лист 'Завод план …' и возвращает:
      { "Вареные": [row_dict, ...], "Ветчины": [...], ... }

    Алгоритм:
      - Строка-разделитель: «замес» == None, «наименование» — название типа → обновляем текущий тип
      - Строка товара: «замес» — число → добавляем в текущую группу
      - Остальные строки (под-заголовки, пустые) → пропускаем
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = _select_target_sheet(wb)

    groups: dict[str, list[dict[str, Any]]] = {}
    current_type: str | None = None

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_slice = row[:_MAX_COLUMNS]

        if all(v is None for v in row_slice):
            continue

        type_name = _detect_type(row_slice)
        if type_name is not None:
            current_type = type_name
            groups.setdefault(current_type, [])
            log.info("Строка %d: тип → '%s'", row_num, current_type)
            continue

        if _is_product_row(row_slice):
            if current_type is None:
                log.warning("Строка %d: товар до первого типа — пропущен: %s", row_num, row_slice)
                continue
            groups[current_type].append({name: row_slice[i] for i, name in _OUTPUT_COLUMNS.items()})
            continue

        log.debug("Строка %d: пропущена (под-заголовок): %s", row_num, row_slice)

    for t, rows in groups.items():
        log.info("Тип '%s': %d товаров", t, len(rows))

    wb.close()
    return groups


def group_rows_by_type_with_days(
    data: bytes,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, dict[str, list[dict[str, Any]]]]]:
    """
    Возвращает:
      - groups: текущая 'потребность' (как раньше)
      - days: { 'ПН 25.03': { 'Вареные': [ {потребность, наименование}, ... ] } }
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = _select_target_sheet(wb)
    day_cols = _detect_day_columns(ws)

    groups: dict[str, list[dict[str, Any]]] = {}
    days: dict[str, dict[str, list[dict[str, Any]]]] = {day_key: {} for day_key in day_cols.values()}
    current_type: str | None = None

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_slice = row[:_MAX_COLUMNS]
        if all(v is None for v in row_slice):
            continue

        type_name = _detect_type(row_slice)
        if type_name is not None:
            current_type = type_name
            groups.setdefault(current_type, [])
            for day_key in days:
                days[day_key].setdefault(current_type, [])
            continue

        if _is_product_row(row_slice):
            if current_type is None:
                log.warning("Строка %d: товар до первого типа — пропущен: %s", row_num, row_slice)
                continue

            base_row = {name: row_slice[i] for i, name in _OUTPUT_COLUMNS.items()}
            groups[current_type].append(base_row)

            # Для каждого дня берем значение конкретной колонки.
            product_name = base_row.get("наименование")
            for col_idx, day_key in day_cols.items():
                day_val = row[col_idx] if col_idx < len(row) else None
                if not isinstance(day_val, (int, float)) or day_val <= 0:
                    continue
                days[day_key][current_type].append(
                    {
                        "потребность": day_val,
                        "наименование": product_name,
                        "замес на 100": base_row.get("замес на 100"),
                        "замес на 150": base_row.get("замес на 150"),
                    }
                )
            continue

    # Чистим пустые типы/дни
    cleaned_days: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for day_key, by_type in days.items():
        cleaned_types = {t: rows for t, rows in by_type.items() if rows}
        if cleaned_types:
            cleaned_days[day_key] = cleaned_types

    wb.close()
    return groups, cleaned_days
